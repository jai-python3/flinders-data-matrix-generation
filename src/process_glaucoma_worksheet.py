# -*- coding: utf-8 -*-
"""Script for processing the DR Excel Worksheet for the Flinders dataset batch
2."""
import json
import logging
import logging.handlers
import os
import pathlib
import sys
import time
from datetime import datetime
from typing import Any, Dict, List, Union

import click
from colorama import Fore, Style
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# Reference: CI SOP

GLAUCOMA_WORKSHEET_NAME = "Glaucoma"

DATASET_NAME = "Flinders_dataset_batch_2"

column_name_to_column_letter_lookup = {
    "Sample_ID": "A",
    "Gender": "B",
    "Ancestry": "C",
    "Glaucoma.diagnosis": "D",
    "Family History": "E",
    "AgeDx": "F",
    "Age Recruitment": "G",
    "Highest IOP_RE": "H",
    "Highest IOP_LE": "I",
    "Highest IOP": "J",
    "NTG HTG": "K",
    "VCDR_RE": "L",
    "VCDR_LE": "M",
    "Highest.VCDR": "N",
}

column_letter_to_column_name_lookup = {
    val: key for key, val in column_name_to_column_letter_lookup.items()
}


EXPECTED_NTG_HTG_VALUES = (
    "0",
    "1",
    "9",
)

EXPECTED_GLAUCOMA_DIAGNOSIS_VALUES = (
    "ASD",
    "GS",
    "LHON",
    "ODD",
    "PACG",
    "PACG, PXF",
    "PCG",
    "PDS",
    "POAG",
    "POAG, PCG",
    "POAG_suspect",
    "PXF",
    # "Unaffected",
)

FINAL_EXPECTED_GLAUCOMA_DIAGNOSIS_VALUES = (
    "ASD",
    "GS",
    "LHON",
    "ODD",
    "PACG",
    "PACG, PXF",
    "PCG",
    "PDS",
    "POAG_strict",  # This includes "POAG" and "POAG, PCG"
    "POAG_loose",  # This is for "POAG_suspect"
    "PXF",
)

EXPECTED_FAMILY_HISTORY_VALUES = (
    "1",
    "No",
    "Not Record",
    "Not Recorded",
    "Unknown",
    "Unsure",
    "Yes",
)


EXPECTED_ANCESTRY_VALUES = (
    "Unknown",
    "Caucasian",
    "Australian",
    "Australian Aboriginal",
    "Middle Eastern",
    "African",
    "Asian",
    "Caucasian/Maori",
    "Hispanic",
    "Lebanese",
    "Melanesian",
    "Mixed Ethnicity",
    "Greek",
    "Egyptian",
    "Polynesian",
    "Ukranian",
    "NA",
)


start_time = time.perf_counter()

MATRIX_YES_VALUE = "2"  # case
MATRIX_NO_VALUE = "1"  # control
MATRIX_NA_VALUE = "NA"

MATRIX_CASE_VALUE = "2"
MATRIX_CONTROL_VALUE = "1"
MATRIX_CASE_CONTROL_NA_VALUE = "0"


MATRIX_GENDER_FEMALE = "2"
MATRIX_GENDER_MALE = "1"
MATRIX_GENDER_NA = "0"


SPLIT_DIAGNOSIS = False

SPLIT_CONTROL_CASE = False

OVERRIDE_CONTROL_CASE = True


DEFAULT_OUTDIR = os.path.join(
    "/tmp/",
    "matrix-generation",
    "flinders",
    os.path.basename(__file__),
    str(datetime.today().strftime("%Y-%m-%d-%H%M%S")),
)

DEFAULT_CONFIG_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "conf/config.json"
)

CONFIG: Dict[str, Any] = {}

DEFAULT_VERBOSE = True

LOGFILE_MAX_BYTES = 50_000
LOGFILE_BACKUP_COUNT = 10

# Set the root logger
logging.basicConfig(format="%(levelname)-7s : %(message)s", level=logging.INFO)

logger = logging.getLogger(__name__)


def setup_filehandler_logger(logfile: str = None):

    # Create handlers
    # c_handler = logging.StreamHandler()
    f_handler = logging.handlers.RotatingFileHandler(
        filename=logfile,
        # maxBytes=LOGFILE_MAX_BYTES,
        # backupCount=LOGFILE_BACKUP_COUNT
    )

    # c_handler.setLevel(logging.INFO)
    f_handler.setLevel(logging.INFO)

    # Create formatters and add it to handlers
    f_format = logging.Formatter(
        "%(levelname)-7s : %(asctime)s : %(pathname)s : L%(lineno)d : %(message)s"
    )
    # c_format = logging.Formatter("%(levelname)-7s : %(asctime)s : %(message)s")

    # c_handler.setFormatter(c_format)
    f_handler.setFormatter(f_format)

    # Add handlers to the logger
    # logger.addHandler(c_handler)
    logger.addHandler(f_handler)


def process_glaucoma_worksheet(sheet_name: str, worksheet, outdir: str) -> None:
    """Process the Glaucoma worksheet.

    Args:
        sheet_name: (str) the name of the worksheet
        worksheet: (Openpyxl Worksheet) the Openpyxl Worksheet object
        outdir: (str) the output directory

    Returns:
        None

    Raises:
        None
    """

    binary_id_lookup: Dict[str, Dict[str, str]] = {}
    quantitative_id_lookup: Dict[str, Dict[str, Union[int, float, str]]] = {}

    # Keep track of the order the Sample_ID values were in the AMD worksheet in the Excel file
    ordered_sample_id_list = []

    row_ctr = 0

    reported_ignored_column_name_lookup = {}

    for row in worksheet:
        row_ctr += 1
        if row_ctr == 1:
            continue  # Skip the header row

        # process this non-header row
        current_sample_id = None

        # Need to retain value of Highest IOP_RE and IOP_LE
        # so that can calculate the mean
        highest_iop_re = None
        highest_iop_le = None

        vcdr_re = None
        vcdr_le = None

        for cell in row:

            cell_value = cell.value
            column_letter = cell.column_letter

            if column_letter not in column_letter_to_column_name_lookup:
                # TODO: Need to log each unique empty column at least once
                # logger.warning(f"Encountered a column letter '{column_letter}' not found in t he column_letter_to_column_name_lookup - will skip it")
                continue

            column_name = column_letter_to_column_name_lookup[column_letter]
            column_name = column_name.strip()  # remove all surrounding whitespace

            if column_name is None:
                logger.error(
                    f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'"
                )
                print_red(
                    f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'"
                )
                sys.exit(1)

            elif column_name in CONFIG["ignore_column_lookup"][sheet_name]:
                if column_name not in reported_ignored_column_name_lookup:
                    logger.info(
                        f"As per configuration setting, will ignore column '{column_name}' in worksheet '{sheet_name}'"
                    )
                    reported_ignored_column_name_lookup[column_name] = True

            elif column_name == "Sample_ID":
                current_sample_id = cell_value
                if current_sample_id is None or current_sample_id == "":
                    logger.warning(
                        f"Found Sample_ID with no value at row '{row_ctr}' in worksheet '{sheet_name}'"
                    )
                    break

                # Initialize the binary lookup for the current sample_id
                if current_sample_id not in binary_id_lookup:
                    binary_id_lookup[current_sample_id] = {}

                # Initialize the quantitative lookup for the current sample_id
                if current_sample_id not in quantitative_id_lookup:
                    quantitative_id_lookup[current_sample_id] = {}

                ordered_sample_id_list.append(current_sample_id)

            elif column_name == "Gender":

                out_column_name = column_name.lower()

                if (
                    cell_value is not None
                    and cell_value != ""
                    and cell_value.lower() != "na"
                    and cell_value.lower() != "unknown"
                ):
                    if cell_value.lower().startswith("m"):
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_GENDER_MALE
                    elif cell_value.lower().startswith("f"):
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_GENDER_FEMALE
                    else:
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE
                else:
                    # print_red(
                    #     f"Encountered unexpected Gender value '{cell_value}' - assigning '{MATRIX_NA_VALUE}'"
                    # )
                    logger.info(
                        f"Encountered unexpected Gender value '{cell_value}' - assigning '{MATRIX_NA_VALUE}'"
                    )

                    binary_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif column_name.strip() == "Ancestry":
                for race in EXPECTED_ANCESTRY_VALUES:
                    out_column_name = f"{column_name.strip().lower()}_{race.lower()}"
                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value.lower() != "na"
                        and cell_value.lower() != "unknown"
                        and cell_value.lower() == race.lower()
                    ):

                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CASE_VALUE
                    else:
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE

            elif column_name == "Glaucoma.diagnosis":

                if cell_value.strip() == "Unaffected":

                    for diagnosis in FINAL_EXPECTED_GLAUCOMA_DIAGNOSIS_VALUES:

                        out_column_name = f"glaucoma_diagnosis_{diagnosis.lower().replace('.', '_').replace(',', '_').replace(' ', '_')}"
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CONTROL_VALUE

                else:
                    for diagnosis in EXPECTED_GLAUCOMA_DIAGNOSIS_VALUES:

                        out_column_name = None
                        final_value = None

                        if diagnosis == "POAG" or diagnosis == "POAG, PCG":
                            out_column_name = f"{column_name.lower().replace('.', '_').replace(',', '_').replace(' ', '_')}_POAG_strict"

                            if (
                                cell_value is not None
                                and cell_value != ""
                                and cell_value.lower() != "na"
                                and cell_value.lower() != "unknown"
                                and (cell_value == "POAG" or cell_value == "POAG, PCG")
                            ):
                                final_value = MATRIX_CASE_VALUE
                            else:
                                final_value = MATRIX_CONTROL_VALUE

                        elif diagnosis == "POAG_suspect":
                            out_column_name = f"{column_name.lower().replace('.', '_').replace(',', '_').replace(' ', '_')}_POAG_loose"
                            if (
                                cell_value is not None
                                and cell_value != ""
                                and cell_value.lower() != "na"
                                and cell_value.lower() != "unknown"
                                and cell_value == "POAG_suspect"
                            ):
                                final_value = MATRIX_CASE_VALUE
                            else:
                                final_value = MATRIX_CONTROL_VALUE

                        else:

                            cell_value = (
                                cell_value.replace(".", "_")
                                .replace(" ", "_")
                                .replace(",", "_")
                            )

                            out_column_name = f"{column_name.lower().replace('.', '_').replace(',', '_').replace(' ', '_')}_{diagnosis.lower().replace('.', '_').replace(',', '_').replace(' ', '_')}"

                            if (
                                cell_value is not None
                                and cell_value != ""
                                and cell_value.lower() != "na"
                                and cell_value.lower() != "unknown"
                                and diagnosis.lower() == cell_value.lower()
                            ):

                                final_value = MATRIX_CASE_VALUE
                            else:
                                final_value = MATRIX_CONTROL_VALUE

                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = final_value

            elif column_name == "Family History":

                for family_history in EXPECTED_FAMILY_HISTORY_VALUES:

                    out_column_name = f"{column_name.lower().replace(' ', '_')}_{family_history.lower().replace(' ', '_')}"

                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value.lower() != "na"
                        and cell_value.lower() != "unknown"
                        and family_history.lower() == cell_value.lower()
                    ):
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CASE_VALUE
                    else:
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE

            elif column_name == "AgeDx":
                # TODO: Need to add support for the non-integer values

                out_column_name = "age_diagnosis"

                if (
                    cell_value is not None
                    and cell_value != ""
                    # and cell_value.lower() != "na"
                    # and cell_value.lower() != "unknown"
                ):
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = cell_value
                else:
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif column_name == "Age Recruitment":
                # TODO: Need to add support for blanks, NA and other spurious values
                out_column_name = column_name.lower().replace(" ", "_")

                if (
                    cell_value is not None
                    and cell_value != ""
                    # and cell_value.lower() != "na"
                    # and cell_value.lower() != "unknown"
                ):
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = cell_value
                else:
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif (
                column_name == "Highest IOP_RE"
                or column_name == "Highest IOP_LE"
                or column_name == "Highest IOP"
            ):

                out_column_name = column_name.lower().replace(" ", "_")

                if type(cell_value) == str:
                    cell_value = cell_value.strip()  # Remove surrounding whitespace
                    if cell_value.lower() == "x":
                        quantitative_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE
                        continue

                if (
                    cell_value is not None
                    and cell_value != ""
                    # and cell_value.lower() != "na"
                    # and cell_value.lower() != "unknown"
                ):
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = cell_value

                    if column_name == "Highest IOP_RE":
                        highest_iop_re = cell_value
                        logger.info(
                            f"Found highest_iop_re '{highest_iop_re}' for sample_id '{current_sample_id}'"
                        )
                    elif column_name == "Highest IOP_LE":
                        highest_iop_le = cell_value
                        logger.info(
                            f"Found highest_iop_le '{highest_iop_le}' for sample_id '{current_sample_id}'"
                        )
                    elif column_name == "Highest IOP":
                        derive_highest_iop_mean(
                            highest_iop_re,
                            highest_iop_le,
                            current_sample_id,
                            quantitative_id_lookup,
                        )

                else:
                    derive_highest_iop_mean(
                        highest_iop_re,
                        highest_iop_le,
                        current_sample_id,
                        quantitative_id_lookup,
                    )
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif column_name == "NTG HTG":

                if type(cell_value) == str:
                    cell_value = cell_value.strip()  # Remove surrounding whitespace

                out_column_name = "high_tension_glaucoma"

                if (
                    cell_value is not None
                    and cell_value != ""
                    # and cell_value.lower() != "na"
                    # and cell_value.lower() != "unknown"
                ):
                    if cell_value == 1:  # High tension glaucoma
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CASE_VALUE
                    elif cell_value == 0:  # Normal tension glaucoma
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CONTROL_VALUE
                    elif cell_value == 9:  #  Not available or not recorded
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE
                else:
                    binary_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif column_name == "VCDR_RE" or column_name == "VCDR_LE":

                if column_name == "VCDR_RE":
                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value != "x"
                        and cell_value != "X"
                    ):
                        if type(cell_value) == str and "-" in cell_value:
                            vcdr_re = get_mean(cell_value)
                        else:
                            vcdr_re = cell_value
                            logger.info(
                                f"Found vcdr_re '{vcdr_re}' for sample_id '{current_sample_id}'"
                            )
                elif column_name == "VCDR_LE":
                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value != "x"
                        and cell_value != "X"
                    ):
                        if type(cell_value) == str and "-" in cell_value:
                            vcdr_re = get_mean(cell_value)
                        else:
                            vcdr_le = cell_value
                            logger.info(
                                f"Found vcdr_le '{vcdr_le}' for sample_id '{current_sample_id}'"
                            )

                if type(cell_value) == str:
                    cell_value = cell_value.strip()  # Remove surrounding whitespace

                out_column_name = column_name.lower().replace(" ", "_")

                if (
                    cell_value is not None
                    and cell_value != ""
                    # and cell_value.lower() != "na"
                    # and cell_value.lower() != "unknown"
                ):
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = cell_value
                else:
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif column_name == "Highest.VCDR":

                derive_vcdr_mean(
                    vcdr_re, vcdr_le, current_sample_id, quantitative_id_lookup
                )

                if type(cell_value) == str:
                    cell_value = cell_value.strip()  # Remove surrounding whitespace

                out_column_name = column_name.lower().replace(".", "_")

                if (
                    cell_value is not None
                    and cell_value != ""
                    # and cell_value.lower() != "na"
                    # and cell_value.lower() != "unknown"
                ):
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = cell_value
                else:
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            else:
                print_red(f"Encountered unexpected column name '{column_name}'")
                logger.warning(f"Encountered unexpected column name '{column_name}'")

    generate_binary_matrix(
        ordered_sample_id_list,
        binary_id_lookup,
        sheet_name,
        f"{os.path.join(outdir, DATASET_NAME + '_' + sheet_name.lower().replace(' ', '_'))}_binary.txt",
    )
    generate_quantitative_matrix(
        ordered_sample_id_list,
        quantitative_id_lookup,
        sheet_name,
        f"{os.path.join(outdir, DATASET_NAME + '_' + sheet_name.lower().replace(' ', '_'))}_quantitative.txt",
    )

    logger.info(f"Processed '{row_ctr}' rows in worksheet '{sheet_name}'")


def get_mean(val: str) -> float:
    x, y = val.split("-")
    mean = (float(x) + float(y)) / 2
    return mean


def derive_vcdr_mean(vcdr_re, vcdr_le, current_sample_id, quantitative_id_lookup):
    """Derive the VCDR mean.

    Args:
        vcdr_re (int): VCDR_RE
        vcdr_le (int): VCDR_LE
        current_sample_id (str): The current sample identifier
        quantitative_id_lookup (dict): The quantitative id lookup

    Returns:
        None
    """
    if vcdr_le is not None and vcdr_re is not None:
        avg = (float(vcdr_le) + float(vcdr_re)) / 2
        logger.info(
            f"Calculated VCDR mean for sample_id '{current_sample_id}': '{avg}'"
        )
        quantitative_id_lookup[current_sample_id]["vcdr_mean"] = avg
    else:
        quantitative_id_lookup[current_sample_id]["vcdr_mean"] = MATRIX_NA_VALUE

        logger.info(
            f"Unable to calculate VCDR mean because vcdr_le is '{vcdr_le}' and vcdr_re is '{vcdr_re}' for sample_id '{current_sample_id}'"
        )


def derive_highest_iop_mean(
    highest_iop_re, highest_iop_le, current_sample_id, quantitative_id_lookup
):
    """Derive the Highest IOP mean.

    Args:
        highest_iop_re (int): Highest IOP_RE
        highest_iop_le (int): Highest IOP_LE
        current_sample_id (str): The current sample identifier
        quantitative_id_lookup (dict): The quantitative id lookup

    Returns:
        None
    """
    if highest_iop_le is not None and highest_iop_re is not None:
        avg = (float(highest_iop_le) + float(highest_iop_re)) / 2
        logger.info(
            f"Calculated highest_iop_mean for sample_id '{current_sample_id}': '{avg}'"
        )
        quantitative_id_lookup[current_sample_id]["highest_iop_mean"] = avg
    else:
        quantitative_id_lookup[current_sample_id]["highest_iop_mean"] = MATRIX_NA_VALUE

        logger.info(
            f"Unable to calculate highest_iop_mean because highest_iop_le is '{highest_iop_le}' and highest_iop_re is '{highest_iop_re}' for sample_id '{current_sample_id}'"
        )


def generate_binary_matrix(
    ordered_sample_id_list: List[str],
    binary_id_lookup: dict,
    sheet_name: str,
    outfile: str,
) -> None:
    """Generate the binary matrix for this worksheet.

    Args:
        binary_id_lookup: (dict) the binary lookup
        sheet_name: (str) the name of the worksheet
        outfile: (str) the output file name

    Returns:
        None

    Raises:
        None
    """

    with open(outfile, "w") as of:
        header_list = []
        header_list.append("ID")

        header_original_list = []
        header_original_list.append("ID")

        ctr = 0

        for sample_id in ordered_sample_id_list:
            if sample_id not in binary_id_lookup:
                print_red(
                    f"Did not find sample_id '{sample_id}' in the binary_id_lookup"
                )
                logger.error(
                    f"Did not find sample_id '{sample_id}' in the binary_id_lookup"
                )

                sys.exit(1)

            ctr += 1
            if ctr == 1:
                for column_name in binary_id_lookup[sample_id]:

                    header_original_list.append(column_name)

                    if (
                        column_name
                        in CONFIG["column_name_conversion_lookup"][sheet_name]
                    ):
                        cn = CONFIG["column_name_conversion_lookup"][sheet_name][
                            column_name
                        ]
                        column_name = cn

                    header_list.append(
                        column_name.lower().replace(" ", "_").replace("-", "_")
                    )

                header_row_str = "\t".join(header_list)
                of.write(f"{header_row_str}\n")

            output_list = []

            for column_name in header_original_list:
                if column_name == "ID":
                    continue
                if column_name not in binary_id_lookup[sample_id]:

                    if (
                        column_name.lower() == "glaucoma_diagnosis_poag_strict"
                        or column_name.lower() == "glaucoma_diagnosis_poag_loose"
                    ):
                        output_list.append(
                            str(binary_id_lookup[sample_id][column_name.lower()])
                        )
                    else:
                        output_list.append(str(MATRIX_NA_VALUE))
                else:
                    output_list.append(str(binary_id_lookup[sample_id][column_name]))

            output_row_str = "\t".join(output_list)
            of.write(f"{sample_id}\t{output_row_str}\n")

        logger.info(f"Wrote '{ctr}' lines to output file '{outfile}'")


def generate_quantitative_matrix(
    ordered_sample_id_list: List[str],
    quantitative_id_lookup: dict,
    sheet_name: str,
    outfile: str,
) -> None:
    """Generate the quantitative matrix for this worksheet.

    Args:
        quantitative_id_matrix: (dict) the quantitative lookup with key being the current sample ID
        sheet_name: (str) the name of the worksheet
        outfile: (str) the output file name

    Returns:
        None

    Raises:
        None
    """
    with open(outfile, "w") as of:
        header_list = []
        header_list.append("ID")

        header_original_list = []
        header_original_list.append("ID")

        ctr = 0

        for sample_id in ordered_sample_id_list:
            if sample_id not in quantitative_id_lookup:
                print_red(
                    f"Did not find sample_id '{sample_id}' in quantitative_id_lookup"
                )
                logger.error(
                    f"Did not find sample_id '{sample_id}' in quantitative_id_lookup"
                )
                sys.exit(1)

            ctr += 1
            if ctr == 1:
                for column_name in quantitative_id_lookup[sample_id]:

                    header_original_list.append(column_name)

                    if (
                        column_name
                        in CONFIG["column_name_conversion_lookup"][sheet_name]
                    ):
                        cn = CONFIG["column_name_conversion_lookup"][sheet_name][
                            column_name
                        ]
                        column_name = cn
                    header_list.append(
                        column_name.lower().replace(" ", "_").replace("-", "_")
                    )
                header_row_str = "\t".join(header_list)
                of.write(f"{header_row_str}\n")

            output_list = []
            output_list.append(sample_id)

            for column_name in header_original_list:
                if column_name == "ID":
                    continue
                if column_name not in quantitative_id_lookup[sample_id]:
                    output_list.append(str(MATRIX_NA_VALUE))
                else:
                    output_list.append(
                        str(quantitative_id_lookup[sample_id][column_name])
                    )

            output_row_str = "\t".join(output_list)
            of.write(f"{output_row_str}\n")

        logger.info(f"Wrote '{ctr}' lines to output file '{outfile}'")


def print_red(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.RED + msg + Style.RESET_ALL)


def print_green(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.GREEN + msg + Style.RESET_ALL)


def print_yellow(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.YELLOW + msg + Style.RESET_ALL)


@click.command()
@click.option(
    "--verbose",
    is_flag=True,
    help=f"Will print more info to STDOUT - default is '{DEFAULT_VERBOSE}'",
)
@click.option(
    "--outdir",
    help="The default is the current working directory - default is '{DEFAULT_OUTDIR}'",
)
@click.option(
    "--config_file",
    type=click.Path(exists=True),
    help=f"The configuration file for this project - default is '{DEFAULT_CONFIG_FILE}'",
)
@click.option("--logfile", help="The log file")
@click.option("--infile", help="The primary input file")
def main(
    verbose: bool,
    outdir: str,
    config_file: str,
    logfile: str,
    infile: str,
):
    """Process command-line arguments and execute main functionality."""
    error_ctr = 0

    if infile is None:
        print_red("--infile was not specified")
        error_ctr += 1

    if error_ctr > 0:
        sys.exit(1)

    assert isinstance(infile, str)

    if config_file is None:
        config_file = os.getenv("CONFIG_FILE", None)
        if config_file is None:
            config_file = DEFAULT_CONFIG_FILE
            print_yellow(
                f"--config_file was not specified and therefore was set to '{config_file}'"
            )
        else:
            print_yellow(f"Will load configuration from '{config_file}'")

    assert isinstance(config_file, str)

    if outdir is None:
        outdir = DEFAULT_OUTDIR
        print_yellow(f"--outdir was not specified and therefore was set to '{outdir}'")

    assert isinstance(outdir, str)

    if not os.path.exists(outdir):
        pathlib.Path(outdir).mkdir(parents=True, exist_ok=True)

        print_yellow(f"Created output directory '{outdir}'")

    if logfile is None:
        logfile = os.path.join(outdir, f"{os.path.basename(__file__)}.log")
        print_yellow(
            f"--logfile was not specified and therefore was set to '{logfile}'"
        )

    assert isinstance(logfile, str)

    setup_filehandler_logger(logfile)

    if not os.path.isfile(infile):
        print_red(f"'{infile}' is not a file")
        logger.error(f"'{infile}' is not a file")
        sys.exit(1)

    logger.info(f"The input file is '{infile}'")

    # Read the configuration from the JSON file and load into dictionary.
    logger.info(f"Loading configuration from '{config_file}'")

    global CONFIG
    CONFIG = json.loads(open(config_file).read())

    logger.info(f"CONFIG: {CONFIG}")

    workbook = load_workbook(filename=infile, data_only=True)

    workbook = load_workbook(filename=infile, data_only=True)
    for sheet_name in workbook.sheetnames:
        logger.info(f"Found sheet name '{sheet_name}'")
        if sheet_name == GLAUCOMA_WORKSHEET_NAME:
            process_glaucoma_worksheet(sheet_name, workbook[sheet_name], outdir)
        else:
            logger.warning(f"Will not process worksheet named '{sheet_name}'")

    logger.info(
        Fore.GREEN
        + f"Execution of '{os.path.abspath(__file__)}' completed"
        + Style.RESET_ALL
    )

    logger.info(f"The log file is '{logfile}'")
    logger.info(f"Total run time was '{time.perf_counter() - start_time}' seconds")
    sys.exit(0)


if __name__ == "__main__":
    main()
