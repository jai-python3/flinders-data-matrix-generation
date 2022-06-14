# -*- coding: utf-8 -*-
"""Script for processing the DR Excel Worksheet for the Flinders dataset batch
2."""
import json
import logging
import logging.handlers
import os
import pathlib
import sys
import time
from datetime import datetime
from typing import Any, Dict, List, Union

import click
from colorama import Fore, Style
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# Reference: CI SOP

GLAUCOMA_WORKSHEET_NAME = "Glaucoma"

DATASET_NAME = "Flinders_dataset_batch_2"

column_name_to_column_letter_lookup = {
    "Sample_ID": "A",
    "Gender": "B",
    "Ancestry": "C",
    "Glaucoma.diagnosis": "D",
    "Family History": "E",
    "AgeDx": "F",
    "Age Recruitment": "G",
    "Highest IOP_RE": "H",
    "Highest IOP_LE": "I",
    "Highest IOP": "J",
    "NTG HTG": "K",
    "VCDR_RE": "L",
    "VCDR_LE": "M",
    "Highest.VCDR": "N",
}

column_letter_to_column_name_lookup = {
    val: key for key, val in column_name_to_column_letter_lookup.items()
}


EXPECTED_NTG_HTG_VALUES = (
    "0",
    "1",
    "9",
)

EXPECTED_GLAUCOMA_DIAGNOSIS_VALUES = (
    "ASD",
    "GS",
    "LHON",
    "ODD",
    "PACG",
    "PACG, PXF",
    "PCG",
    "PDS",
    "POAG",
    "POAG, PCG",
    "POAG_suspect",
    "PXF",
    # "Unaffected",
)

FINAL_EXPECTED_GLAUCOMA_DIAGNOSIS_VALUES = (
    "ASD",
    "GS",
    "LHON",
    "ODD",
    "PACG",
    # "PACG, PXF",
    "PCG",
    "PDS",
    "POAG_strict",  # This includes "POAG" and "POAG, PCG"
    "POAG_loose",  # This is for "POAG_suspect"
    "PXF",
)

EXPECTED_FAMILY_HISTORY_VALUES = (
    "1",
    "No",
    "Not Record",
    "Not Recorded",
    "Unknown",
    "Unsure",
    "Yes",
)


EXPECTED_ANCESTRY_VALUES = (
    "Unknown",
    "Caucasian",
    "Australian",
    "Australian Aboriginal",
    "Middle Eastern",
    "African",
    "Asian",
    "Caucasian/Maori",
    "Hispanic",
    "Lebanese",
    "Melanesian",
    "Mixed Ethnicity",
    "Greek",
    "Egyptian",
    "Polynesian",
    "Ukranian",
    "NA",
)


start_time = time.perf_counter()

MATRIX_YES_VALUE = "2"  # case
MATRIX_NO_VALUE = "1"  # control
MATRIX_NA_VALUE = "NA"

MATRIX_CASE_VALUE = "2"
MATRIX_CONTROL_VALUE = "1"
MATRIX_CASE_CONTROL_NA_VALUE = "0"


MATRIX_GENDER_FEMALE = "2"
MATRIX_GENDER_MALE = "1"
MATRIX_GENDER_NA = "0"


SPLIT_DIAGNOSIS = False

SPLIT_CONTROL_CASE = False

OVERRIDE_CONTROL_CASE = True


DEFAULT_OUTDIR = os.path.join(
    "/tmp/",
    "matrix-generation",
    "flinders",
    os.path.basename(__file__),
    str(datetime.today().strftime("%Y-%m-%d-%H%M%S")),
)

DEFAULT_CONFIG_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "conf/config.json"
)

CONFIG: Dict[str, Any] = {}

DEFAULT_VERBOSE = True

LOGFILE_MAX_BYTES = 50_000
LOGFILE_BACKUP_COUNT = 10

# Set the root logger
logging.basicConfig(format="%(levelname)-7s : %(message)s", level=logging.INFO)

logger = logging.getLogger(__name__)


def setup_filehandler_logger(logfile: str = None):

    # Create handlers
    # c_handler = logging.StreamHandler()
    f_handler = logging.handlers.RotatingFileHandler(
        filename=logfile,
        # maxBytes=LOGFILE_MAX_BYTES,
        # backupCount=LOGFILE_BACKUP_COUNT
    )

    # c_handler.setLevel(logging.INFO)
    f_handler.setLevel(logging.INFO)

    # Create formatters and add it to handlers
    f_format = logging.Formatter(
        "%(levelname)-7s : %(asctime)s : %(pathname)s : L%(lineno)d : %(message)s"
    )
    # c_format = logging.Formatter("%(levelname)-7s : %(asctime)s : %(message)s")

    # c_handler.setFormatter(c_format)
    f_handler.setFormatter(f_format)

    # Add handlers to the logger
    # logger.addHandler(c_handler)
    logger.addHandler(f_handler)


def process_glaucoma_worksheet(sheet_name: str, worksheet, outdir: str) -> None:
    """Process the Glaucoma worksheet.

    Args:
        sheet_name: (str) the name of the worksheet
        worksheet: (Openpyxl Worksheet) the Openpyxl Worksheet object
        outdir: (str) the output directory

    Returns:
        None

    Raises:
        None
    """

    binary_id_lookup: Dict[str, Dict[str, str]] = {}
    quantitative_id_lookup: Dict[str, Dict[str, Union[int, float, str]]] = {}

    # Keep track of the order the Sample_ID values were in the AMD worksheet in the Excel file
    ordered_sample_id_list = []

    row_ctr = 0

    # Tally the number of sample_id values encountered
    sample_id_ctr = +1

    reported_ignored_column_name_lookup = {}

    patient_ctr = 0
    sample_id_to_patient_ctr = {}

    for row in worksheet:
        row_ctr += 1
        if row_ctr == 1:
            continue  # Skip the header row

        # process this non-header row
        current_sample_id = None

        # Need to retain value of Highest IOP_RE and IOP_LE
        # so that can calculate the mean
        highest_iop_re = None
        highest_iop_le = None

        vcdr_re = None
        vcdr_le = None

        for cell in row:

            cell_value = cell.value
            column_letter = cell.column_letter

            if column_letter not in column_letter_to_column_name_lookup:
                # TODO: Need to log each unique empty column at least once
                # logger.warning(f"Encountered a column letter '{column_letter}' not found in t he column_letter_to_column_name_lookup - will skip it")
                continue

            column_name = column_letter_to_column_name_lookup[column_letter]
            column_name = column_name.strip()  # remove all surrounding whitespace

            if column_name is None:
                logger.error(
                    f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'"
                )
                print_red(
                    f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'"
                )
                sys.exit(1)

            elif column_name in CONFIG["ignore_column_lookup"][sheet_name]:
                if column_name not in reported_ignored_column_name_lookup:
                    logger.info(
                        f"As per configuration setting, will ignore column '{column_name}' in worksheet '{sheet_name}'"
                    )
                    reported_ignored_column_name_lookup[column_name] = True

            elif column_name == "Sample_ID":
                current_sample_id = cell_value
                if current_sample_id is None or current_sample_id == "":
                    logger.warning(
                        f"Found Sample_ID with no value at row '{row_ctr}' in worksheet '{sheet_name}'"
                    )
                    break

                # Tally the number of sample_id values encountered
                sample_id_ctr += 1

                # Initialize the binary lookup for the current sample_id
                if current_sample_id not in binary_id_lookup:
                    binary_id_lookup[current_sample_id] = {}

                # Initialize the quantitative lookup for the current sample_id
                if current_sample_id not in quantitative_id_lookup:
                    quantitative_id_lookup[current_sample_id] = {}

                ordered_sample_id_list.append(current_sample_id)

            elif (
                column_name == "Highest IOP_RE"
                or column_name == "Highest IOP_LE"
                or column_name == "Highest IOP"
            ):

                if (
                    cell_value is not None
                    and cell_value != ""
                    # and cell_value.lower() != "na"
                    # and cell_value.lower() != "unknown"
                ):
                    if current_sample_id not in sample_id_to_patient_ctr:
                        patient_ctr += 1
                        sample_id_to_patient_ctr[current_sample_id] = True

    logger.info(f"Processed '{row_ctr}' rows in worksheet '{sheet_name}'")
    logger.info(f"Found '{sample_id_ctr}' sample_id values")
    logger.info(
        f"Found '{patient_ctr}' patient records with at least one IOP measurement"
    )


def print_red(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.RED + msg + Style.RESET_ALL)


def print_green(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.GREEN + msg + Style.RESET_ALL)


def print_yellow(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.YELLOW + msg + Style.RESET_ALL)


@click.command()
@click.option(
    "--verbose",
    is_flag=True,
    help=f"Will print more info to STDOUT - default is '{DEFAULT_VERBOSE}'",
)
@click.option(
    "--outdir",
    help="The default is the current working directory - default is '{DEFAULT_OUTDIR}'",
)
@click.option(
    "--config_file",
    type=click.Path(exists=True),
    help=f"The configuration file for this project - default is '{DEFAULT_CONFIG_FILE}'",
)
@click.option("--logfile", help="The log file")
@click.option("--infile", help="The primary input file")
def main(
    verbose: bool,
    outdir: str,
    config_file: str,
    logfile: str,
    infile: str,
):
    """Process command-line arguments and execute main functionality."""
    error_ctr = 0

    if infile is None:
        print_red("--infile was not specified")
        error_ctr += 1

    if error_ctr > 0:
        sys.exit(1)

    assert isinstance(infile, str)

    if config_file is None:
        config_file = os.getenv("CONFIG_FILE", None)
        if config_file is None:
            config_file = DEFAULT_CONFIG_FILE
            print_yellow(
                f"--config_file was not specified and therefore was set to '{config_file}'"
            )
        else:
            print_yellow(f"Will load configuration from '{config_file}'")

    assert isinstance(config_file, str)

    if outdir is None:
        outdir = DEFAULT_OUTDIR
        print_yellow(f"--outdir was not specified and therefore was set to '{outdir}'")

    assert isinstance(outdir, str)

    if not os.path.exists(outdir):
        pathlib.Path(outdir).mkdir(parents=True, exist_ok=True)

        print_yellow(f"Created output directory '{outdir}'")

    if logfile is None:
        logfile = os.path.join(outdir, f"{os.path.basename(__file__)}.log")
        print_yellow(
            f"--logfile was not specified and therefore was set to '{logfile}'"
        )

    assert isinstance(logfile, str)

    setup_filehandler_logger(logfile)

    if not os.path.isfile(infile):
        print_red(f"'{infile}' is not a file")
        logger.error(f"'{infile}' is not a file")
        sys.exit(1)

    logger.info(f"The input file is '{infile}'")

    # Read the configuration from the JSON file and load into dictionary.
    logger.info(f"Loading configuration from '{config_file}'")

    global CONFIG
    CONFIG = json.loads(open(config_file).read())

    logger.info(f"CONFIG: {CONFIG}")

    workbook = load_workbook(filename=infile, data_only=True)

    workbook = load_workbook(filename=infile, data_only=True)
    for sheet_name in workbook.sheetnames:
        logger.info(f"Found sheet name '{sheet_name}'")
        if sheet_name == GLAUCOMA_WORKSHEET_NAME:
            process_glaucoma_worksheet(sheet_name, workbook[sheet_name], outdir)
        else:
            logger.warning(f"Will not process worksheet named '{sheet_name}'")

    logger.info(
        Fore.GREEN
        + f"Execution of '{os.path.abspath(__file__)}' completed"
        + Style.RESET_ALL
    )

    logger.info(f"The log file is '{logfile}'")
    logger.info(f"Total run time was '{time.perf_counter() - start_time}' seconds")
    sys.exit(0)


if __name__ == "__main__":
    main()
