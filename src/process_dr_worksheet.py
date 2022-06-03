# -*- coding: utf-8 -*-
"""Script for processing the DR Excel Worksheet for the Flinders dataset batch 2."""
import json
import logging
import logging.handlers
import os
import pathlib
import sys
import time
from datetime import datetime
from typing import Any, Dict, List, Union

import click
from colorama import Fore, Style
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# Reference: CI SOP

column_name_to_column_letter_lookup = {
    "Sample_ID": "A",
    "Gender": "B",
    "Ancestry": "C",
    "Age.Recruitment": "D",
    "Disease Type": "E",
    "Year of DR development": "F",
    "BCVA_OD": "G",
    "BCVA_OS": "H",
    "Retinopathy_OD": "I",
    "Retinopathy_OS": "J",
    "Macular Edema_OD": "K",
    "Macular Edema_OS": "L",
    "Control/Case": "M",
}

column_letter_to_column_name_lookup = {
    val: key for key, val in column_name_to_column_letter_lookup.items()
}


EXPECTED_DISEASE_TYPE_VALUES = (
    "ANZRAG",
    "Type1",
    "Type 1",
    "Type2-NIDDM",
    "Type2-IDDM",
    "NA",
)

EXPECTED_ANCESTRY_VALUES = (
    "Unknown",
    "Caucasian",
    "Australian",
    "Australian Aboriginal",
    "Middle Eastern",
    "African",
    "Asian",
    "Caucasian/Maori",
    "Hispanic",
    "Lebanese",
    "Melanesian",
    "Mixed Ethnicity",
    "NA",
)

EXPECTED_RETINOPATHY_TYPE_VALUES = (
    "PDR",
    "Mild NPDR",
    "Severe NPDR",
    "No DR",
    "Minimal NPDR",
    "Moderate NPDR",
    "Unknown",
)

EXPECTED_MACULAR_EDEMA_TYPE_VALUES = ("Yes", "No", "Unknown")


DR_WORKSHEET_NAME = "DR"

DATASET_NAME = "Flinders_dataset_batch_2"

""" If the following is set to True, we will override
the Control/Case value that was assigned by the collaborator. """
OVERRIDE_CONTROL_CASE = True


start_time = time.perf_counter()

MATRIX_YES_VALUE = "2"  # case
MATRIX_NO_VALUE = "1"  # control
MATRIX_NA_VALUE = "NA"

MATRIX_CASE_VALUE = "2"
MATRIX_CONTROL_VALUE = "1"
MATRIX_CASE_CONTROL_NA_VALUE = "0"


MATRIX_GENDER_FEMALE = "2"
MATRIX_GENDER_MALE = "1"
MATRIX_GENDER_NA = "0"


SPLIT_DIAGNOSIS = False

SPLIT_CONTROL_CASE = False

OVERRIDE_CONTROL_CASE = True


DEFAULT_OUTDIR = os.path.join(
    "/tmp/",
    "matrix-generation",
    "flinders",
    os.path.basename(__file__),
    str(datetime.today().strftime("%Y-%m-%d-%H%M%S")),
)

DEFAULT_CONFIG_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "conf/config.json"
)

CONFIG: Dict[str, Any] = {}

DEFAULT_VERBOSE = True

LOGFILE_MAX_BYTES = 50_000
LOGFILE_BACKUP_COUNT = 10

# Set the root logger
logging.basicConfig(format="%(levelname)-7s : %(message)s", level=logging.INFO)

logger = logging.getLogger(__name__)


def setup_filehandler_logger(logfile: str = None):

    # Create handlers
    # c_handler = logging.StreamHandler()
    f_handler = logging.handlers.RotatingFileHandler(
        filename=logfile,
        # maxBytes=LOGFILE_MAX_BYTES,
        # backupCount=LOGFILE_BACKUP_COUNT
    )

    # c_handler.setLevel(logging.INFO)
    f_handler.setLevel(logging.INFO)

    # Create formatters and add it to handlers
    f_format = logging.Formatter(
        "%(levelname)-7s : %(asctime)s : %(pathname)s : L%(lineno)d : %(message)s"
    )
    # c_format = logging.Formatter("%(levelname)-7s : %(asctime)s : %(message)s")

    # c_handler.setFormatter(c_format)
    f_handler.setFormatter(f_format)

    # Add handlers to the logger
    # logger.addHandler(c_handler)
    logger.addHandler(f_handler)


def process_diagnosis(
    column_name: str,
    cell_value: str,
    binary_id_lookup: dict,
    current_sample_id: str,
) -> None:
    """Process the Diagnosis value and store in the binary lookup.

    Args:
        column_name: (str) the column name
        cell_value: (str) the value for the current cell
        binary_id_lookup: (dict) the binary lookup
        current_sample_id: (str) the current sample ID being processed

    Returns:
        None

    Raises:
        None
    """
    # Note from discussion with Kavita 2022-03-22
    # - unaffected is the control
    # - all others are N/A or 2

    cell_value = str(cell_value)  # Convert to a string
    cell_value = cell_value.strip()  # Remove surrounding whitespace
    if "unaffected" in cell_value.lower():
        binary_id_lookup[current_sample_id][column_name] = MATRIX_CONTROL_VALUE
    else:
        binary_id_lookup[current_sample_id][column_name] = MATRIX_CASE_VALUE


def process_yes_no_column(
    column_name: str,
    current_sample_id: str,
    cell_value,
    binary_id_lookup: dict,
) -> None:
    """Process Yes/No column.

    The function will process the columns that are known to contain a Yes or No value.
    The converted value will be stored in the binary lookup.

    Args:
        column_name: (str) the column name
        sheet_name: (str) the name of the current worksheet
        current_sample_id: (str) the current sample ID being processed
        cell_value: (str) the value for the current cell
        binary_id_lookup: (dict) the binary lookup

    Returns:
        None

    Raises:
        None
    """
    final_column_name = column_name.replace(" ", "_").lower()

    cell_value = str(cell_value)  # Convert to a string
    cell_value = cell_value.strip()  # Remove surrounding whitespace

    val = MATRIX_NA_VALUE

    if cell_value is None or cell_value == "None" or cell_value == "":
        val = MATRIX_NA_VALUE
    elif cell_value.lower() == "no":
        val = MATRIX_NO_VALUE
    elif cell_value.lower() == "yes" or cell_value.lower() == "1":
        val = MATRIX_YES_VALUE
    else:
        val = MATRIX_NA_VALUE

    binary_id_lookup[current_sample_id][final_column_name] = val


def process_dr_disease_type(
    current_sample_id: str,
    cell_value,
    column_name: str,
    column_unique_values_lookup: dict,
    binary_id_lookup: dict,
) -> None:
    """Process the Disease Type column in the DR worksheet.

    This function will process the Disease Type column in the DR worksheet
    and store the derived value in the binary lookup.

    Args:
        current_sample_id: (str) the current sample ID being processed
        cell_value: (str) the value for the current cell
        column_name: (str) the column name
        column_unique_values_lookup: (dict) lookup containing unique column values
        binary_id_lookup: (dict) the binary lookup

    Returns:
        None

    Raises:
        None
    """
    for unique_value in column_unique_values_lookup[column_name]:
        if unique_value == "NA":
            continue
        else:
            disease_type = MATRIX_NO_VALUE
            if cell_value == "NA":
                disease_type = MATRIX_NA_VALUE

            categorical_column_name = f"{column_name}_{unique_value}"
            if unique_value == cell_value:
                binary_id_lookup[current_sample_id][
                    categorical_column_name
                ] = MATRIX_YES_VALUE
            else:
                binary_id_lookup[current_sample_id][
                    categorical_column_name
                ] = disease_type


def process_glaucoma_tension(
    sheet_name: str,
    column_name: str,
    cell_value,
    binary_id_lookup: dict,
    current_sample_id: str,
    row_ctr: int,
) -> None:
    """Process normal/high tension column in the Glaucoma worksheet.

    This function will process the normal/high tension value in the Glaucoma
    worksheet and store the derived value in the binary lookup.

    Args:
        sheet_name: (str) the name of the worksheet
        column_name: (str) the column name
        cell_value: (str) the value for the current cell
        binary_id_lookup: (dict) the binary lookup
        current_sample_id: (str) the current sample ID being processed

    Returns:
        None

    Raises:
        None
    """
    cell_value = str(cell_value)  # Convert to a string
    cell_value = cell_value.strip()  # Remove surrounding whitespace

    normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
    high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE

    if cell_value == "0":
        # Notes from discussion with Kavita 2022-03-21:
        # - everyone with normal tension: 2 (case)
        # - everyone with high tension: N/A
        # - all others unaffected are: 1
        # - blanks: N/A

        normal_tension_glaucoma_instance = MATRIX_CASE_VALUE
        high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE

    elif cell_value == "1":
        # Notes from discussion with Kavita 2022-03-21:
        # - everyone with high tension: 2 (case)
        # - everyone with normal tension: N/A
        # - all others unaffected are: 1
        # - blanks: N/A

        high_tension_glaucoma_instance = MATRIX_CASE_VALUE
        normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE

    elif cell_value == "9":
        high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
        normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
    else:
        if (
            column_name in CONFIG["blank_value_allowed"][sheet_name]
            and CONFIG["blank_value_allowed"][sheet_name][column_name] is True
        ):
            high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
            normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
        else:
            msg = f"Unexpected value for column '{column_name}' '{cell_value}' (processing Sample_ID '{current_sample_id}' at row '{row_ctr}')"
            print_red(msg)
            logger.fatal(msg)
            sys.exit(1)

    binary_id_lookup[current_sample_id][
        "normal_tension_glaucoma"
    ] = normal_tension_glaucoma_instance
    binary_id_lookup[current_sample_id][
        "high_tension_glaucoma"
    ] = high_tension_glaucoma_instance


def process_gender(cell_value, binary_id_lookup, current_sample_id) -> None:
    """Process the gender column and store the value in the binary lookup.

    Args:
        cell_value: (str) the value for the current cell
        binary_id_lookup: (dict) the binary lookup
        current_sample_id: (str) the current sample ID being processed

    Returns:
        None

    Raises:
        None
    """
    cell_value = str(cell_value)  # Convert to a string
    cell_value = cell_value.strip()  # Remove surrounding whitespace

    instance_gender = MATRIX_GENDER_NA

    if cell_value.lower() == "f" or cell_value.lower() == "female":
        instance_gender = MATRIX_GENDER_FEMALE
    elif cell_value.lower() == "m" or cell_value.lower() == "male":
        instance_gender = MATRIX_GENDER_MALE
    else:
        instance_gender = MATRIX_GENDER_NA

    binary_id_lookup[current_sample_id]["gender"] = instance_gender


def get_average(value: str) -> float:
    """Split the value and calculate the mean for the two derived values.

    Example value is 6.5-7.8
    Thus the two values for which the mean should be calculated will be
    6.5 and 7.8.

    Args:
        value: (str) the value that should be processed

    Returns:
        float

    Raises:
        None
    """
    low, high = value.split("-")
    average = (float(high) + float(low)) / 2
    return average


def get_column_unique_values_lookup(
    column_name_to_letter_lookup: dict, sheet_name: str, worksheet
) -> dict:
    """Get all the unique values for categorical columns.

    Args:
        column_name_to_letter_lookup: (dict) lookup with key column name and value being the column letter
        sheet_name: (str) name of the worksheet
        worksheet: (Openpyxl Worksheet) the Openpyxl Worksheet object

    Returns:
        Will return a lookup (dict) containing all the unique column values found in the specific column.

    Raises:
        None
    """
    column_unique_values_lookup: Dict[str, Dict[str, int]] = {}

    for column_name, column_letter in column_name_to_letter_lookup.items():
        if (
            column_name
            in CONFIG["worksheet_name_to_column_name_to_be_split_list"][sheet_name]
        ):
            logger.info(
                f"Column '{column_name}' is a categorical column that should be split"
            )
            if column_name not in column_unique_values_lookup:
                column_unique_values_lookup[column_name] = {}

            r_ctr = 0
            for cell in worksheet[column_letter]:
                r_ctr += 1
                if (
                    r_ctr == 1
                    and CONFIG["worksheet_name_to_has_header_row"][sheet_name]
                ):
                    continue

                if cell.value is None:
                    continue

                cell_value = str(cell.value)
                cell_value = cell_value.strip()  # remove surrounding whitespace

                if cell_value is None or cell_value == "None" or cell_value == "":
                    continue

                if sheet_name == "DR" and column_name == "Disease Type":
                    if sheet_name in CONFIG["qualified_disease_type_lookup"]:
                        if (
                            cell_value
                            not in CONFIG["qualified_disease_type_lookup"][sheet_name]
                        ):
                            if cell_value == "Type 1":
                                cell_value = "Type1"
                                logger.info(f"Changed value to '{cell_value}'")
                            else:
                                logger.warning(
                                    f"Will ignore unqualified value '{cell_value}' in worksheet '{sheet_name}' column '{column_name}' row '{r_ctr}'"
                                )
                                continue

                if cell_value not in column_unique_values_lookup[column_name]:
                    column_unique_values_lookup[column_name][cell_value] = 0
                column_unique_values_lookup[column_name][cell_value] += 1

            report_unique_column_values(column_unique_values_lookup, column_name)

    return column_unique_values_lookup


def report_unique_column_values(
    column_unique_values_lookup: dict, column_name: str
) -> None:
    """Report unique column values.

    This function will print to the log file all of the unique values
    found in a particular sheet for a specific column.

    Args:
        column_unique_values_lookup: (dict) lookup containing all unique values found in a specific column
        column_name: (str) the column name

    Returns:
        None

    Raises:
        None
    """
    unique_count = 0
    unique_list = []
    for unique_value in column_unique_values_lookup[column_name]:
        unique_count += 1
        unique_list.append(str(unique_value))
    logger.info(
        f"Found the following '{unique_count}' unique values for categorical column '{column_name}': {','.join(unique_list)}"
    )


def process_header_row(
    row,
    column_name_to_letter_lookup,
    sheet_name,
    worksheet,
    column_name_to_index_lookup,
    index_to_column_name_lookup,
    column_letter_to_column_name_lookup,
) -> dict:
    """Process the header row in the worksheet and return a lookup.

    Args:
        row: (Openpyxl Row) an instance of the Openpyxl Row class
        sheet_name: (str) the name of the worksheet
        worksheet: (Openpyxl Worksheet) the Openpyxl Worksheet object
        column_name_to_index_lookup: (dict) lookup with key column name and value being the column index
        index_to_column_name_lookup: (dict) lookup with key being column index and value being
        column_letter_to_column_name_lookup: (dict) lookup with key being column letter and value being column name

    Returns:
        A lookup (dict) containing all of the unique values found in the specific column

    Raises:
        None
    """
    for i, cell in enumerate(row):
        column_name = cell.value
        if column_name is not None:
            if column_name in CONFIG["ignore_column_lookup"][sheet_name]:
                logger.info(
                    f"Ignoring column '{column_name}' in worksheet '{sheet_name}'"
                )
                continue

            if (
                column_name
                in CONFIG["worksheet_name_to_qualified_column_name_list"][sheet_name]
            ):
                column_name_to_index_lookup[column_name] = i
                column_name_to_letter_lookup[column_name] = cell.column_letter
                index_to_column_name_lookup[i] = column_name
                column_letter_to_column_name_lookup[cell.column_letter] = column_name
                logger.info(
                    f"Found column name '{column_name} in column '{cell.column_letter}'"
                )
            else:
                msg = f"Encountered unqualified column name '{column_name}' for worksheet '{sheet_name}'"
                print_red(msg)
                logger.fatal(msg)
                sys.exit(1)
        else:
            logger.info(
                f"Ignoring column '{cell.column_letter}' since it has no value "
            )

    # Get unique values for all categorial columns
    return get_column_unique_values_lookup(
        column_name_to_letter_lookup, sheet_name, worksheet
    )


def process_amd_worksheet(sheet_name: str, worksheet, outdir: str) -> None:
    """Process the AMD worksheet.

    Args:
        sheet_name: (str) the name of the worksheet
        worksheet: (Openpyxl Worksheet) the Openpyxl Worksheet object
        outdir: (str) the output directory


    Returns:
        None

    Raises:
        None
    """

    binary_id_lookup: Dict[str, Dict[str, str]] = {}
    quantitative_id_lookup: Dict[str, Dict[str, Union[int, float, str]]] = {}

    # Keep track of the order the Sample_ID values were in the AMD worksheet in the Excel file
    ordered_sample_id_list = []

    row_ctr = 0

    for row in worksheet:
        row_ctr += 1
        if row_ctr == 1:
            continue  # Skip the header row

        retinopathy_od = None
        retinopathy_os = None
        macular_edema_od = None
        macular_edema_os = None

        # process this non-header row
        current_sample_id = None

        for cell in row:

            cell_value = cell.value
            column_letter = cell.column_letter

            if column_letter not in column_letter_to_column_name_lookup:
                # TODO: Need to log each unique empty column at least once
                # logger.warning(f"Encountered a column letter '{column_letter}' not found in t he column_letter_to_column_name_lookup - will skip it")
                continue

            column_name = column_letter_to_column_name_lookup[column_letter]
            column_name = column_name.strip()  # remove all surrounding whitespace

            if column_name is None:
                logger.error(
                    f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'"
                )
                print_red(
                    f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'"
                )
                sys.exit(1)

            # elif column_name in CONFIG["ignore_column_lookup"][sheet_name]:
            #     logger.info(
            #         f"Ignoring column '{column_name}' in worksheet '{sheet_name}'"
            #     )

            elif column_name == "Sample_ID":
                current_sample_id = cell_value
                if current_sample_id is None or current_sample_id == "":
                    logger.warning(
                        f"Found Sample_ID with no value at row '{row_ctr}' in worksheet '{sheet_name}'"
                    )
                    break

                # Initialize the binary lookup for the current sample_id
                if current_sample_id not in binary_id_lookup:
                    binary_id_lookup[current_sample_id] = {}

                # Initialize the quantitative lookup for the current sample_id
                if current_sample_id not in quantitative_id_lookup:
                    quantitative_id_lookup[current_sample_id] = {}

                ordered_sample_id_list.append(current_sample_id)

            elif column_name == "Age.Recruitment":

                out_column_name = column_name.lower().replace(".", "_")

                quantitative_id_lookup[current_sample_id][out_column_name] = cell_value

            elif column_name == "Gender":

                out_column_name = column_name.lower()

                if cell_value.lower().startswith("m"):
                    binary_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_GENDER_MALE
                elif cell_value.lower().startswith("f"):
                    binary_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_GENDER_FEMALE
                else:
                    print_red(f"Encountered unexpected Gender value '{cell_value}'")

            elif column_name.strip() == "Ancestry":
                for race in EXPECTED_ANCESTRY_VALUES:
                    out_column_name = f"{column_name.strip().lower()}_{race.lower()}"
                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value.lower() != "na"
                        and cell_value.lower() != "unknown"
                        and cell_value.lower() == race.lower()
                    ):

                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CASE_VALUE
                    else:
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE

            elif column_name == "Disease Type":

                cell_value = cell_value.replace(" ", "")  #  Remove all whitespace

                for disease_type in EXPECTED_DISEASE_TYPE_VALUES:
                    out_column_name = f"{column_name.lower().replace(' ', '_')}_{disease_type.lower()}"
                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value.lower() != "na"
                        and cell_value.lower() != "unknown"
                        and disease_type.lower() == cell_value.lower()
                    ):
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CASE_VALUE
                    else:
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE

            elif column_name == "Year of DR development":

                out_column_name = column_name.lower().replace(" ", "_")

                if (
                    cell_value is not None
                    and cell_value != ""
                    and cell_value != "NA"
                    and cell_value != "Unknown"
                ):
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = cell_value
                else:
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif column_name == "BCVA_OD" or column_name == "BCVA_OS":

                out_column_name = column_name.lower()

                if (
                    cell_value is not None
                    and cell_value != ""
                    and cell_value.lower() != "na"
                    and cell_value.lower() != "unknown"
                ):
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = cell_value
                else:
                    quantitative_id_lookup[current_sample_id][
                        out_column_name
                    ] = MATRIX_NA_VALUE

            elif column_name == "Retinopathy_OD":
                retinopathy_od = str(cell_value.strip())

            elif column_name == "Retinopathy_OS":
                retinopathy_os = str(cell_value.strip())

            elif column_name == "Macular Edema_OD":
                macular_edema_od = str(cell_value.strip())

            elif column_name == "Macular Edema_OS":
                macular_edema_os = str(cell_value.strip())

            elif column_name == "Retinopathy_OD" or column_name == "Retinopathy_OS":

                cell_value = cell_value.strip()  # Remove surrounding whitespace

                for retinopathy_type in EXPECTED_RETINOPATHY_TYPE_VALUES:

                    out_column_name = f"{column_name.lower().replace(' ', '_')}_{retinopathy_type.lower()}"
                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value.lower() != "na"
                        and cell_value.lower() != "unknown"
                        and retinopathy_type.lower() == cell_value.lower()
                    ):
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CASE_VALUE
                    else:
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE

            elif column_name == "Macular Edema_OD" or column_name == "Macular Edema_OS":

                cell_value = cell_value.strip()  # Remove surrounding whitespace

                for macular_edema_type in EXPECTED_MACULAR_EDEMA_TYPE_VALUES:

                    out_column_name = f"{column_name.lower().replace(' ', '_')}_{macular_edema_type.lower()}"
                    if (
                        cell_value is not None
                        and cell_value != ""
                        and cell_value.lower() != "na"
                        and cell_value.lower() != "unknown"
                        and macular_edema_type.lower() == cell_value.lower()
                    ):
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_CASE_VALUE
                    else:
                        binary_id_lookup[current_sample_id][
                            out_column_name
                        ] = MATRIX_NA_VALUE

            elif column_name == "Control/Case":

                out_column_name = column_name.lower().replace("/", "_")

                case_control = None

                if OVERRIDE_CONTROL_CASE:
                    # We will override their control/case designation using our own rules
                    if (
                        retinopathy_od == "No DR"
                        and retinopathy_os == "No DR"
                        and macular_edema_od == "No"
                        and macular_edema_os == "No"
                    ):
                        case_control = MATRIX_CONTROL_VALUE
                    elif (
                        retinopathy_od is None
                        or retinopathy_od == "Unknown"
                        or retinopathy_od == ""
                        or retinopathy_os is None
                        or retinopathy_os == "Unknown"
                        or retinopathy_os == ""
                        or macular_edema_od is None
                        or macular_edema_od == "Unknown"
                        or macular_edema_od == ""
                        or macular_edema_os is None
                        or macular_edema_os == "Unknown"
                        or macular_edema_os == ""
                    ):
                        case_control = MATRIX_NA_VALUE
                    else:
                        case_control = MATRIX_CASE_VALUE
                else:
                    if cell_value == "0":
                        case_control = MATRIX_CONTROL_VALUE
                    elif cell_value == "1":
                        case_control = MATRIX_CASE_VALUE
                    elif cell_value == "9":
                        case_control = MATRIX_NA_VALUE
                    else:
                        # blank?
                        if (
                            column_name in CONFIG["blank_value_allowed"][sheet_name]
                            and CONFIG["blank_value_allowed"][sheet_name][column_name]
                            is True
                        ):
                            case_control = MATRIX_NA_VALUE
                        else:
                            msg = f"Found unexpected value for column '{column_name}' at row '{row_ctr}' in worksheet '{sheet_name}'"
                            print_red(msg)
                            logger.fatal(msg)
                            sys.exit(1)

                binary_id_lookup[current_sample_id][out_column_name] = case_control

            else:
                print_red(f"Encountered unexpected column name '{column_name}'")

    generate_binary_matrix(
        ordered_sample_id_list,
        binary_id_lookup,
        sheet_name,
        f"{os.path.join(outdir, DATASET_NAME + '_' + sheet_name.lower().replace(' ', '_'))}_binary.txt",
    )
    generate_quantitative_matrix(
        ordered_sample_id_list,
        quantitative_id_lookup,
        sheet_name,
        f"{os.path.join(outdir, DATASET_NAME + '_' + sheet_name.lower().replace(' ', '_'))}_quantitative.txt",
    )

    logger.info(f"Processed '{row_ctr}' rows in worksheet '{sheet_name}'")


def generate_binary_matrix(
    ordered_sample_id_list: List[str],
    binary_id_lookup: dict,
    sheet_name: str,
    outfile: str,
) -> None:
    """Generate the binary matrix for this worksheet.

    Args:
        binary_id_lookup: (dict) the binary lookup
        sheet_name: (str) the name of the worksheet
        outfile: (str) the output file name

    Returns:
        None

    Raises:
        None
    """
    with open(outfile, "w") as of:
        header_list = []
        header_list.append("ID")

        header_original_list = []
        header_original_list.append("ID")

        ctr = 0

        for sample_id in ordered_sample_id_list:
            if sample_id not in binary_id_lookup:
                print_red(
                    f"Did not find sample_id '{sample_id}' in the binary_id_lookup"
                )
                sys.exit(1)

            ctr += 1
            if ctr == 1:
                for column_name in binary_id_lookup[sample_id]:

                    header_original_list.append(column_name)

                    if (
                        column_name
                        in CONFIG["column_name_conversion_lookup"][sheet_name]
                    ):
                        cn = CONFIG["column_name_conversion_lookup"][sheet_name][
                            column_name
                        ]
                        column_name = cn

                    header_list.append(
                        column_name.lower().replace(" ", "_").replace("-", "_")
                    )

                header_row_str = "\t".join(header_list)
                of.write(f"{header_row_str}\n")

            output_list = []

            for column_name in header_original_list:
                if column_name == "ID":
                    continue
                if column_name not in binary_id_lookup[sample_id]:
                    output_list.append(str(MATRIX_NA_VALUE))
                else:
                    output_list.append(str(binary_id_lookup[sample_id][column_name]))

            output_row_str = "\t".join(output_list)
            of.write(f"{sample_id}\t{output_row_str}\n")

        logger.info(f"Wrote '{ctr}' lines to output file '{outfile}'")


def generate_quantitative_matrix(
    ordered_sample_id_list: List[str],
    quantitative_id_lookup: dict,
    sheet_name: str,
    outfile: str,
) -> None:
    """Generate the quantitative matrix for this worksheet.

    Args:
        quantitative_id_matrix: (dict) the quantitative lookup with key being the current sample ID
        sheet_name: (str) the name of the worksheet
        outfile: (str) the output file name

    Returns:
        None

    Raises:
        None
    """
    with open(outfile, "w") as of:
        header_list = []
        header_list.append("ID")

        header_original_list = []
        header_original_list.append("ID")

        ctr = 0

        for sample_id in ordered_sample_id_list:
            if sample_id not in quantitative_id_lookup:
                print_red(
                    f"Did not find sample_id '{sample_id}' in quantitative_id_lookup"
                )
                sys.exit(1)

            ctr += 1
            if ctr == 1:
                for column_name in quantitative_id_lookup[sample_id]:

                    header_original_list.append(column_name)

                    if (
                        column_name
                        in CONFIG["column_name_conversion_lookup"][sheet_name]
                    ):
                        cn = CONFIG["column_name_conversion_lookup"][sheet_name][
                            column_name
                        ]
                        column_name = cn
                    header_list.append(
                        column_name.lower().replace(" ", "_").replace("-", "_")
                    )
                header_row_str = "\t".join(header_list)
                of.write(f"{header_row_str}\n")

            output_list = []
            output_list.append(sample_id)

            for column_name in header_original_list:
                if column_name == "ID":
                    continue
                if column_name not in quantitative_id_lookup[sample_id]:
                    output_list.append(str(MATRIX_NA_VALUE))
                else:
                    output_list.append(
                        str(quantitative_id_lookup[sample_id][column_name])
                    )

            output_row_str = "\t".join(output_list)
            of.write(f"{output_row_str}\n")

        logger.info(f"Wrote '{ctr}' lines to output file '{outfile}'")


def print_red(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.RED + msg + Style.RESET_ALL)


def print_green(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.GREEN + msg + Style.RESET_ALL)


def print_yellow(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.

    Args:
        msg (str) the message to be printed

    Returns:
        None

    Raises:
        None
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.YELLOW + msg + Style.RESET_ALL)


@click.command()
@click.option(
    "--verbose",
    is_flag=True,
    help=f"Will print more info to STDOUT - default is '{DEFAULT_VERBOSE}'",
)
@click.option(
    "--outdir",
    help="The default is the current working directory - default is '{DEFAULT_OUTDIR}'",
)
@click.option(
    "--config_file",
    type=click.Path(exists=True),
    help=f"The configuration file for this project - default is '{DEFAULT_CONFIG_FILE}'",
)
@click.option("--logfile", help="The log file")
@click.option("--infile", help="The primary input file")
def main(
    verbose: bool,
    outdir: str,
    config_file: str,
    logfile: str,
    infile: str,
):
    """Process command-line arguments and execute main functionality."""
    error_ctr = 0

    if infile is None:
        print_red("--infile was not specified")
        error_ctr += 1

    if error_ctr > 0:
        sys.exit(1)

    assert isinstance(infile, str)

    if config_file is None:
        config_file = os.getenv("CONFIG_FILE", None)
        if config_file is None:
            config_file = DEFAULT_CONFIG_FILE
            print_yellow(
                f"--config_file was not specified and therefore was set to '{config_file}'"
            )
        else:
            print_yellow(f"Will load configuration from '{config_file}'")

    assert isinstance(config_file, str)

    if outdir is None:
        outdir = DEFAULT_OUTDIR
        print_yellow(f"--outdir was not specified and therefore was set to '{outdir}'")

    assert isinstance(outdir, str)

    if not os.path.exists(outdir):
        pathlib.Path(outdir).mkdir(parents=True, exist_ok=True)

        print_yellow(f"Created output directory '{outdir}'")

    if logfile is None:
        logfile = os.path.join(outdir, f"{os.path.basename(__file__)}.log")
        print_yellow(
            f"--logfile was not specified and therefore was set to '{logfile}'"
        )

    assert isinstance(logfile, str)

    setup_filehandler_logger(logfile)

    if not os.path.isfile(infile):
        print_red(f"'{infile}' is not a file")
        logger.error(f"'{infile}' is not a file")
        sys.exit(1)

    logger.info(f"The input file is '{infile}'")

    # Read the configuration from the JSON file and load into dictionary.
    logger.info(f"Loading configuration from '{config_file}'")

    global CONFIG
    CONFIG = json.loads(open(config_file).read())

    logger.info(f"CONFIG: {CONFIG}")

    workbook = load_workbook(filename=infile, data_only=True)

    workbook = load_workbook(filename=infile, data_only=True)
    for sheet_name in workbook.sheetnames:
        logger.info(f"Found sheet name '{sheet_name}'")
        if sheet_name == DR_WORKSHEET_NAME:
            process_amd_worksheet(sheet_name, workbook[sheet_name], outdir)
        else:
            logger.warning(f"Will not process worksheet named '{sheet_name}'")

    logger.info(
        Fore.GREEN
        + f"Execution of '{os.path.abspath(__file__)}' completed"
        + Style.RESET_ALL
    )

    logger.info(f"The log file is '{logfile}'")
    logger.info(f"Total run time was '{time.perf_counter() - start_time}' seconds")
    sys.exit(0)


if __name__ == "__main__":
    main()
