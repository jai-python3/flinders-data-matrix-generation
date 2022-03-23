from ctypes import create_unicode_buffer
from enum import unique
import os
import sys
import click
import pathlib
import json
import logging
import calendar
import time
import pathlib

from openpyxl import load_workbook

from colorama import Fore, Style
from datetime import datetime

# Reference: CI SOP

MATRIX_YES_VALUE = 2 # case
MATRIX_NO_VALUE = 1 # control
MATRIX_NA_VALUE = 'NA'

MATRIX_CASE_VALUE = 1
MATRIX_CONTROL_VALUE = 2
MATRIX_CASE_CONTROL_NA_VALUE = 0


MATRIX_GENDER_FEMALE = 1
MATRIX_GENDER_MALE = 2
MATRIX_GENDER_NA = 0


SPLIT_DIAGNOSIS = False

SPLIT_CONTROL_CASE = False

OVERRIDE_CONTROL_CASE = True


DEFAULT_OUTDIR = os.path.join(
    '/tmp/',
    os.path.basename(__file__),
    str(datetime.today().strftime('%Y-%m-%d-%H%M%S'))
)

DEFAULT_CONFIG_FILE = os.path.dirname(os.path.abspath(__file__)) + '/conf/config.json'

CONFIG = {}

LOGGING_FORMAT = "%(levelname)s : %(asctime)s : %(pathname)s : %(lineno)d : %(message)s"

LOG_LEVEL = logging.INFO

DEFAULT_VERBOSE = True


def process_dr_disease_type(current_sample_id: str, cell_value, column_name: str, sheet_name: str, column_unique_values_lookup: dict, binary_id_lookup: dict) -> None:

    for unique_value in column_unique_values_lookup[column_name]:
        if unique_value == 'NA':
            continue
        else:
            disease_type = MATRIX_NO_VALUE
            if cell_value == 'NA':
                disease_type = MATRIX_NA_VALUE

            categorical_column_name = f"{column_name}_{unique_value}"
            if unique_value == cell_value:
                binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_YES_VALUE
            else:
                binary_id_lookup[current_sample_id][categorical_column_name] = disease_type


def process_glaucoma_tension(sheet_name: str, column_name: str, cell_value, binary_id_lookup: dict, current_sample_id: str, row_ctr: int) -> None:

    cell_value = str(cell_value) #  Convert to a string
    cell_value = cell_value.strip() #  Remove surrounding whitespace

    normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
    high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE

    if cell_value == '0':
        # Notes from discussion with Kavita 2022-03-21:
        # - everyone with normal tension: 2 (case)
        # - everyone with high tension: N/A
        # - all others unaffected are: 1
        # - blanks: N/A

        normal_tension_glaucoma_instance = MATRIX_CASE_VALUE
        high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE

    elif cell_value == '1':
        # Notes from discussion with Kavita 2022-03-21:
        # - everyone with high tension: 2 (case)
        # - everyone with normal tension: N/A
        # - all others unaffected are: 1
        # - blanks: N/A

        high_tension_glaucoma_instance = MATRIX_CASE_VALUE
        normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE

    elif cell_value == '9':
        high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
        normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
    else:
        if column_name in CONFIG['blank_value_allowed'][sheet_name] and CONFIG['blank_value_allowed'][sheet_name][column_name] == True:
            high_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
            normal_tension_glaucoma_instance = MATRIX_CASE_CONTROL_NA_VALUE
        else:
            msg = f"Unexpected value for column '{column_name}' '{cell_value}' (processing Sample_ID '{current_sample_id}' at row '{row_ctr}')"
            print_red(msg)
            logging.fatal(msg)
            sys.exit(1)

    binary_id_lookup[current_sample_id]['normal_tension_glaucoma'] = normal_tension_glaucoma_instance
    binary_id_lookup[current_sample_id]['high_tension_glaucoma'] = high_tension_glaucoma_instance
    # print(binary_id_lookup)
    # sys.exit(1)


def process_gender(cell_value, binary_id_lookup, current_sample_id) -> None:

    cell_value = str(cell_value) #  Convert to a string
    cell_value = cell_value.strip() #  Remove surrounding whitespace

    # if 'gender' not in binary_id_lookup[current_sample_id]:
    #     binary_id_lookup[current_sample_id]['gender'] = {}
    instance_gender = MATRIX_GENDER_NA

    if cell_value.lower() == 'f' or cell_value.lower() == 'female':
        instance_gender = MATRIX_GENDER_FEMALE
    elif cell_value.lower() == 'm' or cell_value.lower() == 'male':
        instance_gender = MATRIX_GENDER_MALE
    else:
        instance_gender = MATRIX_GENDER_NA

    binary_id_lookup[current_sample_id]['gender'] = instance_gender

def get_average(value: str) -> float:
    low, high = value.split('-')
    average = (float(high) + float(low)) / 2
    return average

# def parse_glaucoma_worksheet(worksheet) -> None:
def get_column_unique_values_lookup(column_name_to_letter_lookup: dict, sheet_name: str, worksheet) -> dict:
    """Get all the unique values for categorical columns
    :param column_name_to_letter_lookup: {dict}
    :param sheet_name: {str}
    :param worksheet: {Openpyxl Worksheet}
    """

    column_unique_values_lookup = {}

    for column_name, column_letter in column_name_to_letter_lookup.items():
        if column_name in CONFIG['worksheet_name_to_column_name_to_be_split_list'][sheet_name]:
            logging.info(f"Column '{column_name}' is a categorical column that should be split")
            if column_name not in column_unique_values_lookup:
                column_unique_values_lookup[column_name] = {}

            r_ctr = 0
            for cell in worksheet[column_letter]:
                r_ctr += 1
                if r_ctr == 1 and CONFIG['worksheet_name_to_has_header_row'][sheet_name]:
                    continue

                if cell.value is None:
                    continue

                cell_value = str(cell.value)
                cell_value = cell_value.strip() # remove surrounding whitespace

                if cell_value is None or cell_value == 'None' or cell_value == '':
                    continue

                if sheet_name == 'DR' and column_name == 'Disease Type':
                    if sheet_name in CONFIG['qualified_disease_type_lookup']:
                        if cell_value not in CONFIG['qualified_disease_type_lookup'][sheet_name]:
                            if cell_value == 'Type 1':
                                cell_value = 'Type1'
                                logging.info(f"Changed value to '{cell_value}'")
                            else:
                                logging.warning(f"Will ignore unqualified value '{cell_value}' in worksheet '{sheet_name}' column '{column_name}' row '{r_ctr}'")
                                continue

                # logging.info(f"Found column value '{cell_value}'")
                if cell_value not in column_unique_values_lookup[column_name]:
                    column_unique_values_lookup[column_name][cell_value] = 0
                column_unique_values_lookup[column_name][cell_value] += 1

            report_unique_column_values(column_unique_values_lookup, sheet_name, column_name)
            
    return column_unique_values_lookup

def report_unique_column_values(column_unique_values_lookup: dict, sheet_name: str, column_name: str) -> None:
    """Report to the log file all the unique values found in a particular sheet for a specific column
    :param column_unique_values_lookup: {dict}
    :param sheet_name: {str}
    :param column_name: {str}
    """
    unique_count = 0
    unique_list = []
    for unique_value in column_unique_values_lookup[column_name]:
        unique_count += 1
        unique_list.append(str(unique_value))
    logging.info(f"Found the following '{unique_count}' unique values for categorical column '{column_name}': {','.join(unique_list)}")


def process_header_row(row, column_name_to_letter_lookup, sheet_name, worksheet, column_name_to_index_lookup, index_to_column_name_lookup, column_letter_to_column_name_lookup) -> dict:
    for i, cell in enumerate(row):
        column_name = cell.value
        if column_name is not None:
            if column_name in CONFIG['ignore_column_lookup'][sheet_name]:
                logging.info(f"Ignoring column '{column_name}' in worksheet '{sheet_name}'")
                continue

            if column_name in CONFIG['worksheet_name_to_qualified_column_name_list'][sheet_name]:
                column_name_to_index_lookup[column_name] = i
                column_name_to_letter_lookup[column_name] = cell.column_letter
                index_to_column_name_lookup[i] = column_name
                column_letter_to_column_name_lookup[cell.column_letter] = column_name
                logging.info(f"Found column name '{column_name} in column '{cell.column_letter}'")
            # print(f"\"{cell.value}\",")
            else:
                msg = f"Encountered unqualified column name '{column_name}' for worksheet '{sheet_name}'"
                print_red(msg)
                logging.fatal(msg)
                sys.exit(1)
        else:
            logging.info(f"Ignoring column '{cell.column_letter}' since it has no value ")
    
    # Get unique values for all categorial columns
    return get_column_unique_values_lookup(column_name_to_letter_lookup, sheet_name, worksheet)


def process_glaucoma_worksheet(sheet_name: str, worksheet, outdir: str) -> None:
    """Process the Glaucoma worksheet
    : sheet_name {str}: The name of the worksheet
    : worksheet {Worksheet}: The openpyxl Worksheet object
    """
    row_ctr = 0
    binary_id_lookup = {}
    quantitative_id_lookup = {}
    column_name_to_index_lookup = {}
    column_name_to_letter_lookup = {}
    column_letter_to_column_name_lookup = {}
    index_to_column_name_lookup = {}


    for row in worksheet:

        current_highest_iop_re = None
        current_vcdr_re = None

        retinopathy_od = None
        retinopathy_os = None
        macular_edema_od = None
        macular_edema_os = None

        row_ctr += 1

        if row_ctr == 1 and CONFIG['worksheet_name_to_has_header_row'][sheet_name]:
            logging.info(f"Found header row in row '{row_ctr}' - will process now")
            column_unique_values_lookup = process_header_row(
                row, 
                column_name_to_letter_lookup, 
                sheet_name, 
                worksheet,
                column_name_to_index_lookup, 
                index_to_column_name_lookup, 
                column_letter_to_column_name_lookup
            )


        else:
            # process this non-header row
            current_sample_id = None

            for cell in row:
                
                cell_value = cell.value
                column_letter = cell.column_letter
                
                if column_letter not in column_letter_to_column_name_lookup:
                    # TODO: Need to log each unique empty column at least once
                    # logging.warning(f"Encountered a column letter '{column_letter}' not found in t he column_letter_to_column_name_lookup - will skip it")
                    continue

                column_name = column_letter_to_column_name_lookup[column_letter]
                column_name = column_name.strip() # remove all surrounding whitespace

                if column_name is None:
                    logging.error(f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'")
                    print_red(f"Encountered column with no name at column letter '{column_letter}' in row '{row_ctr}' in worksheet '{sheet_name}'")
                    sys.exit(1)

                elif column_name in CONFIG['ignore_column_lookup'][sheet_name]:
                    logging.info(f"Ignoring column '{column_name}' in worksheet '{sheet_name}'")
                    continue
                    
                elif column_name == 'Sample_ID':
                    current_sample_id = cell_value
                    if current_sample_id is None or current_sample_id == '':
                        logging.warning(f"Found Sample_ID with no value at row '{row_ctr}' in worksheet '{sheet_name}'")
                        break

                    # Initialize the binary lookup for the current sample_id
                    if current_sample_id not in binary_id_lookup:
                        binary_id_lookup[current_sample_id] = {}

                    # Initialize the quantitative lookup for the current sample_id
                    if current_sample_id not in quantitative_id_lookup:
                        quantitative_id_lookup[current_sample_id] = {}

                
                elif column_name == 'Retinopathy_OD' and sheet_name == 'DR':
                    retinopathy_od = str(cell_value.strip())

                elif column_name == 'Retinopathy_OS' and sheet_name == 'DR':
                    retinopathy_os = str(cell_value.strip())

                elif column_name == 'Macular Edema_OD' and sheet_name == 'DR':
                    macular_edema_od = str(cell_value.strip())

                elif column_name == 'Macular Edema_OS' and sheet_name == 'DR':
                    macular_edema_os = str(cell_value.strip())
                
                elif column_name == 'Control/Case' and sheet_name == 'DR':
                    
                    cell_value = str(cell_value) #  Convert to a string
                    cell_value = cell_value.strip() #  Remove surrounding whitespace

                    if SPLIT_CONTROL_CASE:
                        
                        control_instance = MATRIX_NA_VALUE
                        case_instance = MATRIX_NA_VALUE

                        if cell_value == '0':
                            control_instance = MATRIX_YES_VALUE
                            case_instance = MATRIX_NO_VALUE
                        elif cell_value == '1':
                            case_instance = MATRIX_YES_VALUE
                            control_instance = MATRIX_NO_VALUE
                        elif cell_value == '9':
                            case_instance = MATRIX_NA_VALUE
                            control_instance = MATRIX_NA_VALUE
                        else:
                            if column_name in CONFIG['blank_value_allowed'][sheet_name] and CONFIG['blank_value_allowed'][sheet_name][column_name] == True:
                                continue
                            else:
                                msg = f"Unexpected value for column '{column_name}' '{cell_value}' (processing Sample_ID '{current_sample_id}' at row '{row_ctr}')"
                                print_red(msg)
                                logging.fatal(msg)
                                sys.exit(1)

                        binary_id_lookup[current_sample_id]['control'] = control_instance
                        binary_id_lookup[current_sample_id]['case'] = case_instance
                    else:
                        case_control = None

                        if OVERRIDE_CONTROL_CASE:
                            # We will override their control/case designation using our own rules
                            if retinopathy_od == 'No DR' and retinopathy_os == 'No DR' and macular_edema_od == 'No' and macular_edema_os == 'No':
                                case_control = MATRIX_CONTROL_VALUE
                            elif retinopathy_od is None or \
                                retinopathy_od == 'Unknown' or \
                                retinopathy_od == '' or \
                                retinopathy_os is None or \
                                retinopathy_os == 'Unknown' or \
                                retinopathy_os == '' or \
                                macular_edema_od is None or \
                                macular_edema_od == 'Unknown' or \
                                macular_edema_od == '' or \
                                macular_edema_os is None or \
                                macular_edema_os == 'Unknown' or \
                                macular_edema_os == '':
                                case_control = MATRIX_NA_VALUE
                            else:
                                case_control = MATRIX_CASE_VALUE
                        else:
                            if cell_value == '0':
                                # control
                                case_control = MATRIX_CONTROL_VALUE
                            elif cell_value == '1':
                                # case
                                case_control = MATRIX_CASE_VALUE
                            elif cell_value == '9':
                                # NA
                                case_control = MATRIX_NA_VALUE
                            else:
                                # blank?
                                if column_name in CONFIG['blank_value_allowed'][sheet_name] and CONFIG['blank_value_allowed'][sheet_name][column_name] == True:
                                    case_control = MATRIX_NA_VALUE
                                else:
                                    msg = f"Found blank value for column '{column_name}' at row '{row_ctr}' in worksheet '{sheet_name}'"
                                    print_red(msg)
                                    logging.fatal(msg)
                                    sys.exit(1)
                        binary_id_lookup[current_sample_id][column_name] = case_control

                elif (SPLIT_DIAGNOSIS == False and ((sheet_name == 'Glaucoma' and column_name == 'Glaucoma.diagnosis') or (sheet_name == 'AMD' and column_name == 'Diagnosis'))):
                    # Note from discussion with Kavita 2022-03-22
                    # - unaffected is the control
                    # - all others are N/A or 2

                    cell_value = str(cell_value) #  Convert to a string
                    cell_value = cell_value.strip() #  Remove surrounding whitespace
                    if 'unaffected' in cell_value.lower():                                    
                        binary_id_lookup[current_sample_id][column_name] = MATRIX_CONTROL_VALUE
                    else:
                        binary_id_lookup[current_sample_id][column_name] = MATRIX_CASE_VALUE

                elif column_name == 'Gender':
                    process_gender(cell_value, binary_id_lookup, current_sample_id)

                # elif column_name == 'Family History':
                #     cell_value = str(cell_value) #  Convert to a string
                #     cell_value = cell_value.strip() #  Remove surrounding whitespace

                #     family_history_instance = MATRIX_NA_VALUE

                #     if cell_value.lower() == 'no':
                #         family_history_instance = MATRIX_NO_VALUE
                #     elif cell_value.lower() == 'yes' or  cell_value.lower() == '1':
                #         family_history_instance = MATRIX_YES_VALUE
                #     else:
                #         family_history_instance = MATRIX_NA_VALUE

                #     binary_id_lookup[current_sample_id]['family_history'] = family_history_instance

                elif sheet_name == 'Glaucoma' and column_name == 'NTG HTG':
                    process_glaucoma_tension(sheet_name, column_name, cell_value, binary_id_lookup, current_sample_id, row_ctr)

                else:

                    if column_name in CONFIG['worksheet_name_to_column_name_to_be_split_list'][sheet_name]:
                        # Entered section of code where a new column should 
                        # be established for each unique value in this categorical column
                        # and all cell values for those new columns in that new row 
                        # should be set to NO except to for the one that corresponds with the
                        # actual cell value- for that one, the value should be set to YES.

                        cell_value = str(cell_value) #  Convert to a string
                        cell_value = cell_value.strip() #  Remove surrounding whitespace

                        if column_name == 'Disease Type' and sheet_name == 'DR':
                            if sheet_name in CONFIG['qualified_disease_type_lookup']:
                                if cell_value not in CONFIG['qualified_disease_type_lookup'][sheet_name]:
                                    if cell_value == 'Type 1':
                                        cell_value = 'Type1'
                                        logging.info(f"Changed value to '{cell_value}'")
                                    else:
                                        logging.warning(f"Will ignore unqualified value '{cell_value}' in worksheet '{sheet_name}' column '{column_name}' row '{row_ctr}'")
                                        continue

                        if column_name.lower() == 'diagnosis' and SPLIT_DIAGNOSIS:
                            for unique_value in column_unique_values_lookup[column_name]:
                                if 'unaffected' in cell_value.lower():                                    
                                    if 'unaffected' in unique_value.lower():
                                        continue
                                    else:
                                        categorical_column_name = f"{column_name}_{unique_value}"
                                        binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_NO_VALUE
                                else:
                                    if 'unaffected' in unique_value.lower():
                                        continue
                                    else:
                                        categorical_column_name = f"{column_name}_{unique_value}"
                                        binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_NO_VALUE
                                        if unique_value == cell_value:
                                            binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_YES_VALUE


                                # if unique_value == cell_value:
                                #     binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_YES_VALUE
                                # else:
                                #     categorical_column_name = f"{column_name}_{unique_value}"
                                #     binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_NO_VALUE
                                #     if unique_value == cell_value:
                                #         binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_YES_VALUE
                        elif column_name == 'Disease Type' and sheet_name == 'DR':
                            process_dr_disease_type(current_sample_id, cell_value, column_name, sheet_name, column_unique_values_lookup, binary_id_lookup)

                                # else:
                                #     if current_sample_id == 'FDR0008':
                                #         print(f"Found disease type '{cell_value}' with unique column value '{unique_value}'")
                                #     if unique_value == 'NA':
                                #         continue
                                #     else:
                                #         categorical_column_name = f"{column_name}_{unique_value}"
                                #         binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_NO_VALUE
                                #         if unique_value == cell_value:
                                #             binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_YES_VALUE
                        else:

                            # print(f"Processing column name '{column_name}'")
                            # print(column_unique_values_lookup)
                            # sys.exit(0)
                            for unique_value in column_unique_values_lookup[column_name]:
                                categorical_column_name = f"{column_name}_{unique_value}"
                                binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_NO_VALUE
                                if unique_value == cell_value:
                                    binary_id_lookup[current_sample_id][categorical_column_name] = MATRIX_YES_VALUE

                    elif column_name in CONFIG['worksheet_name_to_column_name_to_be_quantitative_values_list'][sheet_name]:

                        cell_value = str(cell_value) #  Convert to a string
                        cell_value = cell_value.strip() #  Remove surrounding whitespace

                        # Entered section of code where some quantitative value is to be processed
                        final_column_name = column_name.replace(' ', '_').lower()

                        if cell_value is None or cell_value == 'None' or cell_value == '':
                            quantitative_id_lookup[current_sample_id][final_column_name] = MATRIX_NA_VALUE
                            if sheet_name == 'Glaucoma':
                                if column_name == 'Highest IOP_RE' or column_name == 'Highest IOP_LE':
                                    quantitative_id_lookup[current_sample_id]['Highest_IOP_Mean'] = MATRIX_NA_VALUE
                                elif column_name == 'VCDR_RE' or column_name == 'VCDR_LE':
                                    quantitative_id_lookup[current_sample_id]['VCDR_Mean'] = MATRIX_NA_VALUE
                        else:
                            
                            cell_value = cell_value.replace(' ', '') #  remove all whitespace

                            quantitative_id_lookup[current_sample_id][final_column_name] = cell_value

                            if sheet_name == 'Glaucoma':
                                if column_name == 'Highest IOP_RE':
                                    if cell_value.lower() != 'x':
                                        current_highest_iop_re = float(cell_value)
                                elif column_name == 'Highest IOP_LE':
                                    if current_highest_iop_re is None:
                                        quantitative_id_lookup[current_sample_id]['Highest_IOP_Mean'] = MATRIX_NA_VALUE
                                    else:
                                        if cell_value.lower() != 'x':
                                            mean_highest_iop = (float(cell_value) + current_highest_iop_re) / 2
                                            quantitative_id_lookup[current_sample_id]['Highest_IOP_Mean'] = mean_highest_iop
                                        else:
                                            quantitative_id_lookup[current_sample_id]['Highest_IOP_Mean'] = MATRIX_NA_VALUE
                                elif column_name == 'VCDR_RE':
                                    if cell_value.lower() != 'x':
                                        if '-' in cell_value:
                                            # Need to parse and get mean of the range specified e.g.: 0.8-0.9
                                            current_vcdr_re = get_average(cell_value)
                                        else:                                        
                                            current_vcdr_re = float(cell_value)
                                elif column_name == 'VCDR_LE':
                                    if current_vcdr_re is None:
                                        quantitative_id_lookup[current_sample_id]['VCDR_Mean'] = MATRIX_NA_VALUE
                                    else:
                                        if cell_value.lower() != 'x':
                                            vcdr_le = cell_value
                                            if '-' in cell_value:
                                                # Need to parse and get mean of the range specified e.g.: 0.8-0.9
                                                vcdr_le = get_average(cell_value)
                                            mean_vcdr = (float(vcdr_le) + current_vcdr_re) / 2
                                            quantitative_id_lookup[current_sample_id]['VCDR_Mean'] = mean_vcdr
                                        else:
                                            quantitative_id_lookup[current_sample_id]['VCDR_Mean'] = MATRIX_NA_VALUE



                    elif column_name in CONFIG['worksheet_name_to_column_name_yes_no'][sheet_name]:

                        final_column_name = column_name.replace(' ', '_').lower()

                        cell_value = str(cell_value) #  Convert to a string
                        cell_value = cell_value.strip() #  Remove surrounding whitespace

                        val = MATRIX_NA_VALUE

                        if cell_value is None or cell_value == 'None' or cell_value == '':
                            val = MATRIX_NA_VALUE
                        elif cell_value.lower() == 'no':
                            val = MATRIX_NO_VALUE
                        elif cell_value.lower() == 'yes' or  cell_value.lower() == '1':
                            val = MATRIX_YES_VALUE
                        else:
                            val = MATRIX_NA_VALUE

                        binary_id_lookup[current_sample_id][final_column_name] = val

                    else:
                        logging.fatal(f"Unexpected column '{column_name}' at row '{row_ctr}' in sheet '{sheet_name}'")
                        print_red(f"Unexpected column '{column_name}' at row '{row_ctr}' in sheet '{sheet_name}'")
                        sys.exit(1)

    # print(binary_id_lookup)
    # sys.exit(1)

    generate_binary_matrix(binary_id_lookup, sheet_name, f"{os.path.join(outdir, sheet_name.lower().replace(' ', '_'))}_binary.txt")
    generate_quantitative_matrix(quantitative_id_lookup, sheet_name, f"{os.path.join(outdir, sheet_name.lower().replace(' ', '_'))}_quantitative.txt")

    print(f"Processed '{row_ctr}' rows in worksheet '{sheet_name}'")


def generate_binary_matrix(binary_id_lookup: dict, sheet_name: str, outfile: str) -> None:
    """Generate the binary matrix for this worksheet
    :binary_id_lookup {dict}: 
    :sheet_name {str}: the worksheet name
    :outfile {str}: the output file path
    """
    
    with open(outfile, 'w') as of:
        header_list = []
        header_list.append('ID')
        ctr = 0

        for sample_id in binary_id_lookup:
            # print(f"Processing sample_id '{sample_id}'")
            ctr += 1
            if ctr == 1:
                for column_name in binary_id_lookup[sample_id]:
                    header_list.append(column_name)
                header_row_str = "\t".join(header_list) 
                of.write(f"{header_row_str}\n")
                # print(f"header row string: {header_row_str}")

            output_list = []
            # output_list.append(sample_id)
            for column_name in header_list:
                # print(f"Processing column name '{column_name}'")
                if column_name == 'ID':
                    continue
                if column_name not in binary_id_lookup[sample_id]:
                    output_list.append(str(MATRIX_NA_VALUE))
                else:
                    output_list.append(str(binary_id_lookup[sample_id][column_name]))
            # print(output_list)

            output_row_str = "\t".join(output_list)
            of.write(f"{sample_id}\t{output_row_str}\n")

        print(f"Wrote '{ctr}' lines to output file '{outfile}'")
        logging.info(f"Wrote '{ctr}' lines to output file '{outfile}'")

    
def generate_quantitative_matrix(quantitative_id_lookup:  dict, sheet_name: str, outfile: str) -> None:
    """Generate the quantitative matrix for this worksheet
    :quantitative_id_matrix {dict}: 
    :sheet_name {str}: the worksheet name
    :outfile {str}: the output file path
    """
    with open(outfile, 'w') as of:
        header_list = []
        header_list.append('ID')
        ctr = 0

        for sample_id in quantitative_id_lookup:
            ctr += 1
            if ctr == 1:
                for column_name in quantitative_id_lookup[sample_id]:
                    header_list.append(column_name)
                header_row_str = "\t".join(header_list) 
                of.write(f"{header_row_str}\n")
            
            output_list = []
            output_list.append(sample_id)
            for column_name in header_list:
                if column_name == 'ID':
                    continue
                if column_name not in quantitative_id_lookup[sample_id]:
                    output_list.append(str(MATRIX_NA_VALUE))
                else:
                    output_list.append(str(quantitative_id_lookup[sample_id][column_name]))
            
            output_row_str = "\t".join(output_list)
            of.write(f"{output_row_str}\n")
        
        print(f"Wrote '{ctr}' lines to output file '{outfile}'")
        logging.info(f"Wrote '{ctr}' lines to output file '{outfile}'")


def print_red(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.
    :param msg: {str} - the message to be printed
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.RED + msg)
    print(Style.RESET_ALL + "", end="")


def print_green(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.
    :param msg: {str} - the message to be printed
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.GREEN + msg)
    print(Style.RESET_ALL + "", end="")


def print_yellow(msg: str = None) -> None:
    """Print message to STDOUT in yellow text.
    :param msg: {str} - the message to be printed
    """
    if msg is None:
        raise Exception("msg was not defined")

    print(Fore.YELLOW + msg)
    print(Style.RESET_ALL + "", end="")


@click.command()
@click.option('--verbose', is_flag=True, help=f"Will print more info to STDOUT - default is '{DEFAULT_VERBOSE}'")
@click.option('--outdir', help="The default is the current working directory - default is '{DEFAULT_OUTDIR}'")
@click.option('--config_file', type=click.Path(exists=True), help=f"The configuration file for this project - default is '{DEFAULT_CONFIG_FILE}'")
@click.option('--logfile', help="The log file")
@click.option('--outfile', help="The output final report file")
@click.option('--infile', help="The primary input file")
def main(verbose: bool, outdir: str, config_file: str, logfile: str, outfile: str, infile: str):
    """Template command-line executable
    """

    error_ctr = 0

    if infile is None:
        print_red("--infile was not specified")
        error_ctr += 1

    if error_ctr > 0:
        sys.exit(1)

    assert isinstance(infile, str)

    if config_file is None:
        config_file = DEFAULT_CONFIG_FILE
        print_yellow(f"--config_file was not specified and therefore was set to '{config_file}'")

    assert isinstance(config_file, str)

    if outdir is None:
        outdir = DEFAULT_OUTDIR
        print_yellow(f"--outdir was not specified and therefore was set to '{outdir}'")

    assert isinstance(outdir, str)

    if not os.path.exists(outdir):
        pathlib.Path(outdir).mkdir(parents=True, exist_ok=True)

        print_yellow(f"Created output directory '{outdir}'")

    if logfile is None:
        logfile = os.path.join(
            outdir,
            f"{os.path.basename(__file__)}.log"
        )
        print_yellow(f"--logfile was not specified and therefore was set to '{logfile}'")

    assert isinstance(logfile, str)

    logging.basicConfig(filename=logfile, format=LOGGING_FORMAT, level=LOG_LEVEL)

    if not os.path.isfile(infile):
        print(f"'{infile}' is not a file")
        logging.error(f"'{infile}' is not a file")
        sys.exit(1)

    if verbose:
        print(f"The input file is '{infile}'")

    logging.info(f"The input file is '{infile}'")

    # Read the configuration from the JSON file and
    # load into dictionary.
    logging.info(f"Loading configuration from '{config_file}'")

    global CONFIG
    CONFIG = json.loads(open(config_file).read())

    logging.info(f"CONFIG: {CONFIG}")

    workbook = load_workbook(filename=infile, data_only=True)
    for sheet_name in workbook.sheetnames:
        logging.info(f"Found sheet name '{sheet_name}'")
        if sheet_name in CONFIG['qualified_sheet_names']:
            logging.info(f"Found qualified sheet named '{sheet_name}'")
            if sheet_name in CONFIG['sheets_to_process']:
                logging.info(f"Will process work sheet '{sheet_name}'")
                if sheet_name == 'Glaucoma' or sheet_name == 'DR':
                    process_glaucoma_worksheet(sheet_name, workbook[sheet_name], outdir)
                    # pass
                else:
                    process_glaucoma_worksheet(sheet_name, workbook[sheet_name], outdir)
            else:
                logging.warning(f"Will not process worksheet named '{sheet_name}'")
        else:
            logging.warning(f"Found unqualified sheet named '{sheet_name}'")
        # sheet = workbook[sheet_name]
        # print(f"The title of the Worksheet is: {sheet.title}")
        # print(f"Cells that contain data: {sheet.calculate_dimension()}")

if __name__ == "__main__":
    main()


